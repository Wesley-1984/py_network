#!/usr/local/bin/python3
# coding:utf-8
# 作者:浮尘
import netmiko
import paramiko
import time
import pandas as pd
from openpyxl import load_workbook
from netmiko import ConnectHandler


class BackupConfig(object):
    def __init__(self):
        """初始化参数"""
        self.device_file = '/Users/admin/Desktop/交换机巡检模板.xlsx'

    def load_excel(self):
        """加载excel文件"""
        try:
            wb = load_workbook(self.device_file)
            return wb
        except FileNotFoundError:
            print(f"{self.device_file}文件不存在")

    def get_device_info(self):
        """获取设备的基础信息：IP地址/用户名/密码等"""
        try:
            # # 方法1：openpyxl
            # wb = self.load_excel()
            # ws1 = wb[wb.sheetnames[0]]
            # # 通过参数min_row,max_col限制区域
            # for row in ws1.iter_rows(min_row=2, max_col=9):
            #     if str(row[1].value).strip() == '#':
            #         # 跳过注释行
            #         continue
            #     info_dict = {'ip': row[2].value,
            #                  'protocol': row[3].value,
            #                  'prot': row[4].value,
            #                  'username': row[5].value,
            #                  'password': row[6].value,
            #                  'secret': row[7].value,
            #                  'device_type': row[8].value,
            #                  'cmd_list': self.get_cmd_info(wb[row[8].value.strip().lower()])
            #                  }
            #     yield info_dict

            # 方法2：pandas
            names = ['comment', 'ip', 'protocol', 'port', 'username', 'password', 'secret', 'device_type']
            df = pd.read_excel(self.device_file, usecols='B:I', names=names, keep_default_na=False)
            data = df.to_dict(orient='records')
            for row in data:
                if row['comment'] == '#':
                    continue
                row['cmd_list'] = self.get_cmd_info(row['device_type'])
                yield row
        except Exception as e:
            print("Error:", e)
        # finally:
        #     # 记得最后关闭workbook
        #     wb.close()

    def get_cmd_info(self, cmd_sheet):
        """获取命令的信息，返回所有命令的列表"""

        cmd_list = []
        try:
            # # 方法1：openpyxl
            # for row in cmd_sheet.iter_rows(min_row=2, max_col=2):
            #     if str(row[0].value).strip() != "#" and row[1].value:
            #         # 跳过注释行，支掉命令左右空格
            #         cmd_list.append(row[1].value.strip())
            # return cmd_list
            names = ['comment', 'cmd']
            df = pd.read_excel(self.device_file, sheet_name=cmd_sheet, usecols='A:B', names=names,
                               keep_default_na=False)
            data = df.to_dict(orient='records')
            for row in data:
                if row['comment'].strip() != '#' and row['cmd']:
                    cmd_list.append(row['cmd'].strip())
            return cmd_list
        except Exception as e:
            print("get_cmd_info Error: ", e)

    def connectHandler(self, host):
        """定义一个netmiko对象"""
        try:
            connect = ''
            # 判断使用ssh协议
            if host['protocol'].lower().strip() == 'ssh':
                host['port'] = host['port'] if (host['port'] not in [22, None]) else 22
                host.pop('protocol'), host.pop('cmd_list'), host.pop('comment')
                # 判断华为有些设备登录超时
                if 'huawei' in host['device_type']:
                    connect = ConnectHandler(**host, conn_timeout=15)
                else:
                    connect = ConnectHandler(**host)
            # 判断使用telnet协议
            elif host['protocol'].lower().strip() == 'telnet':
                host['port'] = host['port'] if (host['port'] not in [23, None]) else 23
                host.pop('protocol'), host.pop('cmd_list')
                # netmiko  使用telnet device_type需要加_telnet
                host['device_type'] = host['device_type'] + '_telnet'
                connect = ConnectHandler(**host, fast_cli=False)
            # 不支持的协议
            else:
                res = f"{host['ip']}暂不支持{host['protocol']}协议"
                raise ValueError(res)
            return connect

        except netmiko.NetmikoAuthenticationException:
            print(host['ip'], '认证失败，请检查用户名和密码')
        except netmiko.ConnectionException:
            print(host['ip'], '连接失败', )
        except netmiko.NetmikoTimeoutException:
            print(host['ip'], '网络不可达')

    def run_cmd(self, host, cmds, enable=False):
        """执行命令和保存信息"""
        # 特权功能标识位
        enable = True if host['secret'] else False
        try:
            conn = self.connectHandler(host)
            if conn:
                # 获取到设备名称，不同人或不同厂商命令都会不同，按需优化
                hostname = conn.find_prompt()
                # if cmds:
                #     output = ''
                for cmd in cmds:
                    if enable:
                        # 进入特权模式
                        conn.enable()
                        output = conn.send_command(cmd, strip_command=False, strip_prompt=False)
                        print('=' * 30, hostname, '=' * 30)
                        print(output)
                    else:
                        output = conn.send_command(cmd, strip_command=False, strip_prompt=False)
                        print(hostname.center(80,'='))
                        print('输入命令：', output)
                else:
                    conn.disconnect()
        except Exception as e:
            print(f'run_cmd error{e}')

    def connect(self):
        pass

    def main(self):
        """主程序"""
        # hosts是一个生成器，需要for循环进行遍历
        hosts = self.get_device_info()
        for host in hosts:
            self.run_cmd(host, host['cmd_list'])
        print('所有已经执行完成')


if __name__ == "__main__":
    BackupConfig().main()
